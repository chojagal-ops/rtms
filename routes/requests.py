# routes/requests.py — 신뢰성 시험 의뢰서 CRUD + 엑셀 파일 파싱 업로드

import os, json
from io import BytesIO
from datetime import date, datetime
from werkzeug.utils import secure_filename
from flask import (Blueprint, render_template, request, redirect,
                   url_for, flash, current_app, send_from_directory, jsonify, send_file)
from flask_login import login_required, current_user
from models import db, TestRequest, TestItem, TestCriterion, TestStandard
from utils import parse_date, log_error, mail_new_request, mail_accept_notify
from constants import (PRODUCT_TYPES, TEST_STAGES, TEST_PURPOSES,
                       SAMPLE_STATES, PRIORITIES, FEASIBILITY_OPTIONS,
                       CRITERION_TYPES, ATTACH_TYPES)

requests_bp = Blueprint('requests', __name__)

UPLOAD_FOLDER = 'static/uploads/requests'
ALLOWED_EXT   = {'pdf', 'png', 'jpg', 'jpeg', 'xlsx', 'xls', 'docx', 'doc', 'zip', 'hwp'}


def _allowed(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXT


def _save_file(file):
    """파일 저장 후 저장된 파일명 반환"""
    if not file or file.filename == '':
        return None
    if not _allowed(file.filename):
        flash('허용되지 않는 파일 형식입니다.', 'warning')
        return None
    from datetime import datetime
    upload_dir = os.path.join(current_app.root_path, UPLOAD_FOLDER)
    os.makedirs(upload_dir, exist_ok=True)
    prefix = datetime.now().strftime('%Y%m%d%H%M%S_')
    fname  = prefix + secure_filename(file.filename)
    file.save(os.path.join(upload_dir, fname))
    return fname


def _next_request_no():
    """의뢰번호 자동 채번 (RTN-YYYYMM-NNN)"""
    today = date.today()
    prefix = f'RTN-{today.strftime("%Y%m")}'
    last = (TestRequest.query
            .filter(TestRequest.request_no.like(f'{prefix}%'))
            .order_by(TestRequest.request_no.desc())
            .first())
    seq = 1
    if last and last.request_no:
        try:
            seq = int(last.request_no.split('-')[-1]) + 1
        except (ValueError, IndexError):
            seq = 1
    return f'{prefix}-{seq:03d}'


def _parse_excel_request(filepath):
    """
    신뢰성 시험 의뢰서 엑셀 파일을 파싱해서 dict 반환.
    COM 없이 openpyxl 만 사용 → 수식 결과값은 data_only=True 로 읽음
    """
    import openpyxl
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active
    except Exception:
        return None

    def cell(coord):
        """셀 값을 문자열로 반환 (None → '')"""
        v = ws[coord].value
        return str(v).strip() if v is not None else ''

    data = {
        'write_date':          cell('C4') or cell('D4'),
        'deadline':            cell('F4'),
        'request_dept':        cell('C6'),
        'requester_name':      cell('C7'),
        'contact':             cell('C9'),
        'project_customer':    cell('C10'),
        'product_name':        cell('F6'),
        'model_code':          cell('F7'),
        'product_type':        cell('G8'),
        'test_stage':          cell('G9'),
        'change_content':      cell('F10'),
        'test_purpose':        cell('D12'),
        'test_purpose_detail': cell('C13'),
        'sample_qty':          cell('F12'),
        'sample_state':        cell('G13'),
        'sample_id_method':    cell('F14'),
        'sample_notes':        cell('F15'),
        'req_start_date':      cell('C29'),
        'req_end_date':        cell('C30'),
        'priority':            cell('D31'),
        'schedule_notes':      cell('C32'),
        'other_conditions':    cell('F32'),
        'attach_doc_name':     cell('C35'),
        'writer_name':         cell('C37'),
        'dept_approver':       cell('C38'),
        # 시험 항목 (행 18~21, 열 B~F)
        'items': [],
    }

    # 시험 항목 파싱 (행 18~21)
    for row_no, row_idx in enumerate(range(18, 22), 1):
        name  = cell(f'C{row_idx}')
        cond  = cell(f'D{row_idx}')
        std   = cell(f'F{row_idx}')
        if name or cond or std:
            data['items'].append({
                'item_no': row_no,
                'test_name': name,
                'test_condition': cond,
                'standard': std,
            })

    return data


def _form_to_request(req_obj, form, files=None):
    """폼 데이터를 TestRequest 객체에 매핑"""
    req_obj.write_date          = parse_date(form.get('write_date'))
    req_obj.deadline            = parse_date(form.get('deadline'))
    req_obj.request_dept        = form.get('request_dept', '').strip()
    req_obj.requester_name      = form.get('requester_name', '').strip()
    req_obj.requester_position  = form.get('requester_position', '').strip()
    req_obj.contact             = form.get('contact', '').strip()
    req_obj.project_customer    = form.get('project_customer', '').strip()
    req_obj.product_name        = form.get('product_name', '').strip()
    req_obj.model_code          = form.get('model_code', '').strip()
    req_obj.product_type        = form.get('product_type', '').strip()
    req_obj.test_stage          = form.get('test_stage', '').strip()
    req_obj.change_content      = form.get('change_content', '').strip()
    req_obj.test_purpose        = form.get('test_purpose', '').strip()
    req_obj.test_purpose_detail = form.get('test_purpose_detail', '').strip()
    qty = form.get('sample_qty', '').strip()
    req_obj.sample_qty          = int(qty) if qty.isdigit() else None
    req_obj.sample_state        = form.get('sample_state', '').strip()
    req_obj.sample_id_method    = form.get('sample_id_method', '').strip()
    req_obj.sample_notes        = form.get('sample_notes', '').strip()
    req_obj.req_start_date      = parse_date(form.get('req_start_date'))
    req_obj.req_end_date        = parse_date(form.get('req_end_date'))
    req_obj.priority            = form.get('priority', '').strip()
    req_obj.schedule_notes      = form.get('schedule_notes', '').strip()
    req_obj.test_unit           = form.get('test_unit') == 'on'
    req_obj.test_package        = form.get('test_package') == 'on'
    req_obj.test_assembly       = form.get('test_assembly') == 'on'
    req_obj.other_conditions    = form.get('other_conditions', '').strip()
    req_obj.attach_types        = ','.join(form.getlist('attach_types'))
    req_obj.attach_doc_name     = form.get('attach_doc_name', '').strip()
    req_obj.writer_name         = form.get('writer_name', '').strip()
    req_obj.dept_approver       = form.get('dept_approver', '').strip()
    req_obj.receiver_name       = form.get('receiver_name', '').strip()
    req_obj.feasibility         = form.get('feasibility', '').strip()
    req_obj.review_opinion      = form.get('review_opinion', '').strip()
    req_obj.qa_approver         = form.get('qa_approver', '').strip()
    req_obj.test_type           = form.get('test_type', '신규시험').strip()
    rid_ref = form.get('retest_ref_id', '').strip()
    req_obj.retest_ref_id       = int(rid_ref) if rid_ref.isdigit() else None
    req_obj.notes               = form.get('notes', '').strip()
    if files:
        saved = _save_file(files.get('attachment'))
        if saved:
            req_obj.attachment = saved


# ── 시험 기준서 API (자동입력용) ────────────────────────────
def _split_condition_criterion(full_text):
    """condition_full 텍스트에서 시험조건과 판정기준을 분리"""
    if not full_text:
        return '', ''
    for marker in ['▶ 판정기준', '◆ 판정기준', '■ 판정기준', '● 판정기준', '※ 판정기준',
                   '►판정기준', '▶판정기준', '▶ 합격기준', '◆ 합격기준', '합격기준\n']:
        idx = full_text.find(marker)
        if idx != -1:
            return full_text[:idx].rstrip(), full_text[idx:]
    # 구분자 없으면 전체를 시험조건으로 반환
    return full_text, ''


@requests_bp.route('/api/test-standard-books')
@login_required
def api_test_standard_books():
    """기준서 목록 API — 모달 북 선택 드롭다운용"""
    rows = db.session.query(TestStandard.book_name).distinct().order_by(TestStandard.book_name).all()
    return jsonify([b[0] for b in rows if b[0]])


@requests_bp.route('/api/test-standards')
@login_required
def api_test_standards():
    """기준서 항목 검색 API — 의뢰서 작성 시 시험항목 자동입력용"""
    q    = request.args.get('q', '').strip()
    book = request.args.get('book', '').strip()
    page = int(request.args.get('page', 1))
    per  = 20  # 한 페이지 항목 수
    query = TestStandard.query
    if book:
        query = query.filter(TestStandard.book_name == book)
    if q:
        query = query.filter(
            db.or_(
                TestStandard.test_name.ilike(f'%{q}%'),
                TestStandard.condition_summary.ilike(f'%{q}%')
            )
        )
    total = query.count()
    items = query.order_by(TestStandard.std_no).offset((page-1)*per).limit(per).all()
    result = []
    for s in items:
        cond, crit = _split_condition_criterion(s.condition_full)
        result.append({
            'id':        s.id,
            'no':        s.std_no,
            'name':      s.test_name,
            'summary':   s.condition_summary,
            'full':      s.condition_full,
            'condition': cond,   # 시험조건만
            'criterion': crit,   # 판정기준만
            'qty':       s.sample_qty,
        })
    return jsonify({'total': total, 'page': page, 'items': result})


# ── 목록 ────────────────────────────────────────────────────
@requests_bp.route('/requests')
@login_required
def list_view():
    status_filter = request.args.get('status', '')
    dept_filter   = request.args.get('dept', '').strip()
    search        = request.args.get('q', '').strip()
    try:
        q = TestRequest.query
        if status_filter:
            q = q.filter_by(status=status_filter)
        if dept_filter:
            q = q.filter(TestRequest.request_dept == dept_filter)
        if search:
            q = q.filter(
                db.or_(
                    TestRequest.request_no.ilike(f'%{search}%'),
                    TestRequest.product_name.ilike(f'%{search}%'),
                    TestRequest.request_dept.ilike(f'%{search}%'),
                    TestRequest.requester_name.ilike(f'%{search}%'),
                )
            )
        items = q.order_by(TestRequest.created_at.desc()).all()

        dept_rows = (db.session.query(TestRequest.request_dept)
                     .distinct()
                     .filter(TestRequest.request_dept.isnot(None))
                     .all())
        depts = sorted([d[0] for d in dept_rows if d[0]])
    except Exception as e:
        log_error('의뢰서 목록 조회 오류', e)
        flash('목록을 불러오는 중 오류가 발생했습니다.', 'danger')
        items = []
        depts = []
    return render_template('requests/list.html', items=items,
                           status_filter=status_filter, dept_filter=dept_filter,
                           search=search, depts=depts)


# ── 신규 등록 (수동 입력) ───────────────────────────────────
@requests_bp.route('/requests/new', methods=['GET', 'POST'])
@login_required
def new():
    prefill = {}   # 엑셀 파싱 결과 pre-fill 데이터
    if request.method == 'POST':
        try:
            req_obj = TestRequest(
                created_by=current_user.id,
                request_no=_next_request_no(),
            )
            _form_to_request(req_obj, request.form, request.files)
            db.session.add(req_obj)
            db.session.flush()   # id 확보

            # 시험 항목 저장
            _save_test_items(req_obj.id, request.form)
            # 판정 기준 저장
            _save_criteria(req_obj.id, request.form)

            db.session.commit()
            flash(f'의뢰서 [{req_obj.request_no}] 등록되었습니다.', 'success')
            # 품질팀 접수 알림 메일 발송 (비동기)
            try:
                base_url = request.url_root.rstrip('/')
                mail_new_request(req_obj, base_url)
            except Exception as e:
                log_error('접수 메일 발송 오류', e)
            return redirect(url_for('requests.detail', rid=req_obj.id))
        except Exception as e:
            db.session.rollback()
            log_error('의뢰서 등록 오류', e)
            flash('등록 중 오류가 발생했습니다.', 'danger')
    prev_reqs = (TestRequest.query
                 .filter(TestRequest.status.in_(['완료', '결과회신']))
                 .order_by(TestRequest.created_at.desc()).all())
    return render_template('requests/form.html', item=None, prefill=prefill,
                           PRODUCT_TYPES=PRODUCT_TYPES, TEST_STAGES=TEST_STAGES,
                           TEST_PURPOSES=TEST_PURPOSES, SAMPLE_STATES=SAMPLE_STATES,
                           PRIORITIES=PRIORITIES, FEASIBILITY_OPTIONS=FEASIBILITY_OPTIONS,
                           CRITERION_TYPES=CRITERION_TYPES, ATTACH_TYPES=ATTACH_TYPES,
                           prev_reqs=prev_reqs)


# ── 엑셀 업로드 파싱 (AJAX) ────────────────────────────────
@requests_bp.route('/requests/parse-excel', methods=['POST'])
@login_required
def parse_excel():
    """의뢰서 엑셀 파일 업로드 → 파싱 결과 JSON 반환"""
    file = request.files.get('excel_file')
    if not file or file.filename == '':
        return jsonify({'ok': False, 'msg': '파일을 선택해주세요.'})
    ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else ''
    if ext not in ('xlsx', 'xls'):
        return jsonify({'ok': False, 'msg': 'xlsx / xls 파일만 지원합니다.'})

    import tempfile
    with tempfile.NamedTemporaryFile(suffix=f'.{ext}', delete=False) as tmp:
        file.save(tmp.name)
        tmp_path = tmp.name
    try:
        data = _parse_excel_request(tmp_path)
        if data is None:
            return jsonify({'ok': False, 'msg': '파일을 읽을 수 없습니다. 엑셀(.xlsx) 형식인지 확인하세요.'})
        return jsonify({'ok': True, 'data': data})
    except Exception as e:
        log_error('엑셀 파싱 오류', e)
        return jsonify({'ok': False, 'msg': '파싱 중 오류가 발생했습니다.'})
    finally:
        try:
            os.remove(tmp_path)
        except Exception:
            pass


# ── 수정 ────────────────────────────────────────────────────
@requests_bp.route('/requests/<int:rid>/edit', methods=['GET', 'POST'])
@login_required
def edit(rid):
    req_obj = db.get_or_404(TestRequest, rid)
    if request.method == 'POST':
        try:
            _form_to_request(req_obj, request.form, request.files)
            # 시험 항목 전체 교체
            TestItem.query.filter_by(request_id=rid).delete()
            _save_test_items(rid, request.form)
            TestCriterion.query.filter_by(request_id=rid).delete()
            _save_criteria(rid, request.form)
            db.session.commit()
            flash('의뢰서가 수정되었습니다.', 'success')
            return redirect(url_for('requests.detail', rid=rid))
        except Exception as e:
            db.session.rollback()
            log_error('의뢰서 수정 오류', e)
            flash('수정 중 오류가 발생했습니다.', 'danger')
    prev_reqs = (TestRequest.query
                 .filter(TestRequest.status.in_(['완료', '결과회신']))
                 .order_by(TestRequest.created_at.desc()).all())
    return render_template('requests/form.html', item=req_obj, prefill={},
                           PRODUCT_TYPES=PRODUCT_TYPES, TEST_STAGES=TEST_STAGES,
                           TEST_PURPOSES=TEST_PURPOSES, SAMPLE_STATES=SAMPLE_STATES,
                           PRIORITIES=PRIORITIES, FEASIBILITY_OPTIONS=FEASIBILITY_OPTIONS,
                           CRITERION_TYPES=CRITERION_TYPES, ATTACH_TYPES=ATTACH_TYPES,
                           prev_reqs=prev_reqs)


# ── 상세 보기 ───────────────────────────────────────────────
@requests_bp.route('/requests/<int:rid>')
@login_required
def detail(rid):
    req_obj = db.get_or_404(TestRequest, rid)
    return render_template('requests/detail.html', item=req_obj)


# ── 삭제 ────────────────────────────────────────────────────
@requests_bp.route('/requests/<int:rid>/delete', methods=['POST'])
@login_required
def delete(rid):
    req_obj = db.get_or_404(TestRequest, rid)
    try:
        db.session.delete(req_obj)
        db.session.commit()
        flash('의뢰서가 삭제되었습니다.', 'success')
    except Exception as e:
        db.session.rollback()
        log_error('의뢰서 삭제 오류', e)
        flash('삭제 중 오류가 발생했습니다.', 'danger')
    return redirect(url_for('requests.list_view'))


# ── 품질팀 접수 처리 ────────────────────────────────────────
@requests_bp.route('/requests/<int:rid>/accept', methods=['POST'])
@login_required
def accept(rid):
    req_obj = db.get_or_404(TestRequest, rid)
    try:
        req_obj.receiver_name   = request.form.get('receiver_name', '').strip()
        req_obj.feasibility     = request.form.get('feasibility', '').strip()
        req_obj.qa_approver     = request.form.get('qa_approver', '').strip()
        req_obj.review_opinion  = request.form.get('review_opinion', '').strip()
        # 접수대기 상태일 때만 접수완료로 자동 전환
        if req_obj.status == '접수대기':
            req_obj.status = '접수완료'
        db.session.commit()
        flash('접수 처리가 완료되었습니다.', 'success')
        # 의뢰자에게 접수 완료 알림 메일 발송 (비동기)
        try:
            from models import User
            requester = User.query.get(req_obj.created_by)
            if requester and requester.email:
                base_url = request.url_root.rstrip('/')
                mail_accept_notify(req_obj, requester.email, base_url)
        except Exception as e:
            log_error('접수완료 알림 메일 발송 오류', e)
    except Exception as e:
        db.session.rollback()
        log_error('접수 처리 오류', e)
        flash('처리 중 오류가 발생했습니다.', 'danger')
    return redirect(url_for('requests.detail', rid=rid))


# ── 시험 항목 복사용 API ────────────────────────────────────
@requests_bp.route('/requests/recent-json')
@login_required
def recent_json():
    q_str = request.args.get('q', '').strip()
    q = TestRequest.query
    if q_str:
        q = q.filter(db.or_(
            TestRequest.request_no.ilike(f'%{q_str}%'),
            TestRequest.product_name.ilike(f'%{q_str}%'),
        ))
    rows = q.order_by(TestRequest.created_at.desc()).limit(20).all()
    return jsonify([{
        'id': r.id,
        'request_no': r.request_no or '',
        'product_name': r.product_name or '',
        'item_count': len(r.test_items),
    } for r in rows])


@requests_bp.route('/requests/<int:rid>/items-json')
@login_required
def items_json(rid):
    req_obj = db.get_or_404(TestRequest, rid)
    return jsonify({
        'request_no': req_obj.request_no,
        'product_name': req_obj.product_name,
        'items': [{'test_name': ti.test_name or '', 'test_condition': ti.test_condition or '', 'standard': ti.standard or ''}
                  for ti in req_obj.test_items],
    })


_REQ_TEMPLATE_NET  = (r'\\192.168.10.3\품질팀\▣ SEC 무선'
                      r'\신뢰성 시험\신뢰성 시험 의뢰서 접수'
                      r'\신뢰성_시험_의뢰서_Rev.0_251113.xlsx')
# Render.com 등 클라우드 환경: static/excel_templates/ 에 파일을 커밋해두면 사용
_REQ_TEMPLATE_LOCAL = os.path.join(
    os.path.dirname(os.path.dirname(__file__)),
    'static', 'excel_templates', '신뢰성_시험_의뢰서_Rev.0_251113.xlsx'
)


def _req_template_path():
    """네트워크 공유 → 로컬 번들 순서로 템플릿 경로 반환. 없으면 None."""
    for p in (_REQ_TEMPLATE_NET, _REQ_TEMPLATE_LOCAL):
        if os.path.isfile(p):
            return p
    return None


# ── 의뢰서 엑셀 내보내기 (원본 양식 기반) ───────────────────
@requests_bp.route('/requests/<int:rid>/export-excel')
@login_required
def export_excel(rid):
    import openpyxl
    from openpyxl.styles import Alignment

    req_obj = db.get_or_404(TestRequest, rid)
    tpl = _req_template_path()
    if not tpl:
        flash('템플릿 파일을 찾을 수 없습니다. '
              'static/excel_templates/ 폴더에 의뢰서 양식 파일을 넣어주세요.', 'danger')
        return redirect(url_for('requests.detail', rid=rid))
    try:
        wb = openpyxl.load_workbook(tpl)
    except Exception as e:
        log_error('의뢰서 템플릿 로드 실패', e)
        flash('템플릿 파일을 불러올 수 없습니다.', 'danger')
        return redirect(url_for('requests.detail', rid=rid))

    ws = wb.active
    lw = Alignment(horizontal='left', vertical='center', wrap_text=True)

    def w(addr, val):
        c = ws[addr]
        c.value = str(val) if val else ''
        c.alignment = lw

    fd = lambda d: d.strftime('%Y-%m-%d') if d else ''

    w('C3', req_obj.request_no)
    w('C4', fd(req_obj.write_date))
    w('F4', fd(req_obj.deadline))
    w('C6', req_obj.request_dept)
    w('C7', f"{req_obj.requester_name or ''} / {req_obj.requester_position or ''}")
    w('C9', req_obj.contact)
    w('C10', req_obj.project_customer)
    w('F6', req_obj.product_name)
    w('F7', req_obj.model_code)
    w('F8', req_obj.product_type)
    w('F9', req_obj.test_stage)
    w('F10', req_obj.change_content)
    w('C12', req_obj.test_purpose)
    w('C13', req_obj.test_purpose_detail)
    w('F12', (str(req_obj.sample_qty) + '개') if req_obj.sample_qty else '')
    w('F13', req_obj.sample_state)
    w('F14', req_obj.sample_id_method)
    w('F15', req_obj.sample_notes)

    for i, row in enumerate(range(18, 22)):
        ti = req_obj.test_items[i] if i < len(req_obj.test_items) else None
        w(f'C{row}', ti.test_name if ti else '')
        w(f'D{row}', ti.test_condition if ti else '')
        w(f'F{row}', ti.standard if ti else '')

    for i, row in enumerate(range(24, 28)):
        cr = req_obj.criteria[i] if i < len(req_obj.criteria) else None
        w(f'C{row}', cr.criterion_type if cr else '')
        w(f'D{row}', cr.criterion_content if cr else '')

    w('C29', fd(req_obj.req_start_date))
    w('C30', fd(req_obj.req_end_date))
    w('C31', req_obj.priority)
    w('C32', req_obj.schedule_notes)
    w('F29', '예' if req_obj.test_unit else '아니오')
    w('F30', '예' if req_obj.test_package else '아니오')
    w('F31', '예' if req_obj.test_assembly else '아니오')
    w('F32', req_obj.other_conditions)
    w('C34', req_obj.attach_types or '')
    w('C35', req_obj.attach_doc_name)
    w('C37', req_obj.writer_name)
    w('C38', req_obj.dept_approver)
    w('C40', req_obj.receiver_name)
    w('C41', req_obj.feasibility)
    w('C42', req_obj.review_opinion)
    w('C43', req_obj.qa_approver)

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    fname = f'신뢰성_시험_의뢰서_{req_obj.request_no or rid}_{datetime.now().strftime("%Y%m%d")}.xlsx'
    return send_file(out, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ── 품질팀 업무함 ────────────────────────────────────────────
@requests_bp.route('/qteam')
@login_required
def qteam():
    """품질팀 업무함 — 상태별 의뢰서 현황 및 워크플로우"""
    from models import TestRequest as TR
    try:
        waiting  = TR.query.filter_by(status='접수대기').order_by(TR.created_at.asc()).all()
        received = TR.query.filter_by(status='접수완료').order_by(TR.created_at.asc()).all()
        testing  = TR.query.filter_by(status='시험중').order_by(TR.created_at.asc()).all()
        replied  = TR.query.filter_by(status='결과회신').order_by(TR.created_at.asc()).all()
        hold     = TR.query.filter_by(status='보류').order_by(TR.created_at.asc()).all()
    except Exception as e:
        log_error('품질팀 업무함 조회 오류', e)
        waiting = received = testing = replied = hold = []
    return render_template('qteam.html',
        waiting=waiting, received=received,
        testing=testing, replied=replied, hold=hold,
    )


# ── 빠른 상태 변경 ───────────────────────────────────────────
@requests_bp.route('/requests/<int:rid>/quick-status', methods=['POST'])
@login_required
def quick_status(rid):
    """품질팀 업무함용 빠른 상태 변경"""
    req_obj = db.get_or_404(TestRequest, rid)
    new_status = request.form.get('status', '').strip()
    valid = {'접수완료', '시험중', '완료', '보류', '결과회신', '접수대기'}
    if new_status not in valid:
        flash('유효하지 않은 상태값입니다.', 'warning')
        return redirect(request.referrer or url_for('requests.qteam'))
    try:
        req_obj.status = new_status
        # 접수완료 처리 시 접수 정보 함께 저장
        if new_status == '접수완료':
            rname = request.form.get('receiver_name', '').strip()
            if rname:
                req_obj.receiver_name = rname
            feas = request.form.get('feasibility', '').strip()
            if feas:
                req_obj.feasibility = feas
            opinion = request.form.get('review_opinion', '').strip()
            if opinion:
                req_obj.review_opinion = opinion
        db.session.commit()
        flash(f'[{req_obj.request_no}] 상태가 [{new_status}]로 변경되었습니다.', 'success')
        # 접수완료 전환 시 의뢰자에게 알림 메일 발송 (비동기)
        if new_status == '접수완료':
            try:
                from models import User
                requester = User.query.get(req_obj.created_by)
                if requester and requester.email:
                    base_url = request.url_root.rstrip('/')
                    mail_accept_notify(req_obj, requester.email, base_url)
            except Exception as e:
                log_error('접수완료 알림 메일 발송 오류', e)
    except Exception as e:
        db.session.rollback()
        log_error('빠른 상태 변경 오류', e)
        flash('변경 중 오류가 발생했습니다.', 'danger')
    return redirect(request.referrer or url_for('requests.qteam'))


# ── 파일 다운로드 ────────────────────────────────────────────
@requests_bp.route('/requests/uploads/<filename>')
@login_required
def download_file(filename):
    upload_dir = os.path.join(current_app.root_path, UPLOAD_FOLDER)
    return send_from_directory(upload_dir, filename, as_attachment=True)


# ── 의뢰서 인쇄 뷰 ───────────────────────────────────────────
@requests_bp.route('/requests/<int:rid>/certificate')
@login_required
def certificate(rid):
    """의뢰서 인쇄용 HTML 뷰"""
    req_obj = db.get_or_404(TestRequest, rid)
    return render_template('requests/certificate.html', item=req_obj)


# ── 의뢰서 엑셀 다운로드 (programmatic) ──────────────────────
@requests_bp.route('/requests/<int:rid>/certificate/excel')
@login_required
def certificate_excel(rid):
    """의뢰서 성적서 엑셀 생성 (openpyxl 직접 생성)"""
    from io import BytesIO
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import datetime as dt

    req_obj = db.get_or_404(TestRequest, rid)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '시험의뢰서'

    col_widths = [6, 18, 28, 14, 14, 24]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    thin  = Side(style='thin',   color='000000')
    thick = Side(style='medium', color='000000')
    b_all = Border(left=thin, right=thin, top=thin, bottom=thin)

    def cell(row, col, value='', bold=False, size=10, align='left',
             valign='center', fill=None, color='000000', wrap=False):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(name='맑은 고딕', bold=bold, size=size, color=color)
        c.alignment = Alignment(horizontal=align, vertical=valign, wrap_text=wrap)
        if fill:
            c.fill = PatternFill('solid', fgColor=fill)
        c.border = b_all
        return c

    def merge(r1, c1, r2, c2):
        ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)

    fd = lambda d: d.strftime('%Y-%m-%d') if d else '-'
    ORANGE = 'F97316'
    LGRAY  = 'F1F5F9'
    DARK   = '334155'
    GRAY   = '475569'

    # ── 제목 ───────────────────────────────────────────────────
    ws.row_dimensions[1].height = 14
    ws.row_dimensions[2].height = 36
    merge(2, 1, 2, 6)
    c = ws.cell(row=2, column=1, value='신뢰성 시험 의뢰서')
    c.font = Font(name='맑은 고딕', bold=True, size=18, color='FFFFFF')
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.fill = PatternFill('solid', fgColor=ORANGE)
    c.border = b_all

    ws.row_dimensions[3].height = 14
    merge(4, 1, 4, 3)
    c = ws.cell(row=4, column=1, value=f'의뢰번호: {req_obj.request_no or "-"}')
    c.font = Font(name='맑은 고딕', bold=True, size=10)
    c.alignment = Alignment(horizontal='left', vertical='center')
    merge(4, 4, 4, 6)
    c = ws.cell(row=4, column=4, value=f'작성일: {fd(req_obj.write_date)}')
    c.font = Font(name='맑은 고딕', size=10)
    c.alignment = Alignment(horizontal='right', vertical='center')
    ws.row_dimensions[4].height = 18
    ws.row_dimensions[5].height = 6

    def section_title(row, title):
        merge(row, 1, row, 6)
        c = ws.cell(row=row, column=1, value=title)
        c.font = Font(name='맑은 고딕', bold=True, size=10, color=GRAY)
        c.alignment = Alignment(horizontal='left', vertical='center')
        c.fill = PatternFill('solid', fgColor=LGRAY)
        c.border = b_all
        ws.row_dimensions[row].height = 16

    def th(row, col, value):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(name='맑은 고딕', bold=True, size=9, color='FFFFFF')
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.fill = PatternFill('solid', fgColor=DARK)
        c.border = b_all
        return c

    def td(row, col, value, bold=False, color='000000', align='center', wrap=False):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(name='맑은 고딕', size=10, bold=bold, color=color)
        c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
        c.border = b_all
        return c

    # ── 1. 의뢰 부서 및 담당자 ───────────────────────────────
    section_title(6, '1. 의뢰 부서 및 담당자')
    ws.row_dimensions[7].height = 16
    ws.row_dimensions[8].height = 18
    for col, h in enumerate(['의뢰부서', '의뢰자', '직책', '연락처', '프로젝트/고객사', '작성일'], 1):
        th(7, col, h)
    vals = [req_obj.request_dept or '-', req_obj.requester_name or '-',
            req_obj.requester_position or '-', req_obj.contact or '-',
            req_obj.project_customer or '-', fd(req_obj.write_date)]
    for col, v in enumerate(vals, 1):
        td(8, col, v)

    ws.row_dimensions[9].height = 6

    # ── 2. 제품 정보 ─────────────────────────────────────────
    section_title(10, '2. 제품 정보')
    ws.row_dimensions[11].height = 16
    ws.row_dimensions[12].height = 18
    for col, h in enumerate(['제품명', '모델명/코드', '제품 종류', '시험 단계', '변경 내용', '완료 희망일'], 1):
        th(11, col, h)
    for col, v in enumerate([req_obj.product_name or '-', req_obj.model_code or '-',
                              req_obj.product_type or '-', req_obj.test_stage or '-',
                              req_obj.change_content or '-', fd(req_obj.deadline)], 1):
        td(12, col, v)

    ws.row_dimensions[13].height = 6

    # ── 3. 시험 목적 및 시편 ─────────────────────────────────
    section_title(14, '3. 시험 목적 및 시편')
    ws.row_dimensions[15].height = 16
    ws.row_dimensions[16].height = 18
    for col, h in enumerate(['시험 목적', '시편 수량', '시편 상태', '식별 방법', '우선순위', '특이사항'], 1):
        th(15, col, h)
    for col, v in enumerate([req_obj.test_purpose or '-',
                              f"{req_obj.sample_qty}개" if req_obj.sample_qty else '-',
                              req_obj.sample_state or '-', req_obj.sample_id_method or '-',
                              req_obj.priority or '-', req_obj.sample_notes or '-'], 1):
        td(16, col, v)

    ws.row_dimensions[17].height = 6

    # ── 4. 시험 항목 ─────────────────────────────────────────
    section_title(18, '4. 시험 항목')
    ws.row_dimensions[19].height = 16
    for col, h in enumerate(['No', '시험 항목', '시험 조건', '규격', '일정 시작', '일정 완료'], 1):
        th(19, col, h)

    row = 20
    items = req_obj.test_items
    for ti in items:
        ws.row_dimensions[row].height = 28
        td(row, 1, ti.item_no, align='center')
        td(row, 2, ti.test_name or '-', bold=True, align='left')
        c = ws.cell(row=row, column=3, value=ti.test_condition or '-')
        c.font = Font(name='맑은 고딕', size=9)
        c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        c.border = b_all
        td(row, 4, ti.standard or '-', align='center')
        td(row, 5, fd(req_obj.req_start_date))
        td(row, 6, fd(req_obj.req_end_date))
        row += 1

    if not items:
        merge(20, 1, 20, 6)
        c = ws.cell(row=20, column=1, value='등록된 시험 항목이 없습니다.')
        c.font = Font(name='맑은 고딕', size=9, color='94A3B8')
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = b_all
        row = 21

    ws.row_dimensions[row].height = 6
    row += 1

    # ── 5. 판정 기준 ─────────────────────────────────────────
    section_title(row, '5. 판정 기준')
    row += 1
    ws.row_dimensions[row].height = 16
    th(row, 1, 'No')
    merge(row, 2, row, 3)
    th(row, 2, '기준 유형')
    merge(row, 4, row, 6)
    th(row, 4, '기준 내용')
    row += 1

    criteria = req_obj.criteria
    for cr in criteria:
        ws.row_dimensions[row].height = 22
        td(row, 1, cr.criterion_no, align='center')
        merge(row, 2, row, 3)
        td(row, 2, cr.criterion_type or '-', align='center')
        merge(row, 4, row, 6)
        c = ws.cell(row=row, column=4, value=cr.criterion_content or '-')
        c.font = Font(name='맑은 고딕', size=10)
        c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        c.border = b_all
        row += 1

    if not criteria:
        merge(row, 1, row, 6)
        c = ws.cell(row=row, column=1, value='등록된 판정 기준이 없습니다.')
        c.font = Font(name='맑은 고딕', size=9, color='94A3B8')
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = b_all
        row += 1

    ws.row_dimensions[row].height = 6
    row += 1

    # ── 6. 품질팀 접수 정보 ─────────────────────────────────
    section_title(row, '6. 품질팀 접수 정보')
    row += 1
    ws.row_dimensions[row].height = 16
    ws.row_dimensions[row+1].height = 18
    for col, h in enumerate(['접수자', '시험 가능 여부', 'QA 승인자', '완료희망일', '작성자', '부서 승인'], 1):
        th(row, col, h)
    feas_color = ('15803d' if req_obj.feasibility == '가능' else
                  'dc2626' if req_obj.feasibility == '불가' else '374151')
    for col, (v, color) in enumerate([
        (req_obj.receiver_name or '-', '000000'),
        (req_obj.feasibility or '-',   feas_color),
        (req_obj.qa_approver or '-',   '000000'),
        (fd(req_obj.deadline),         '000000'),
        (req_obj.writer_name or '-',   '000000'),
        (req_obj.dept_approver or '-', '000000'),
    ], 1):
        c = ws.cell(row=row+1, column=col, value=v)
        c.font = Font(name='맑은 고딕', size=10, bold=(col==2), color=color)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = b_all
    row += 2

    ws.row_dimensions[row].height = 6
    row += 1

    # ── 7. 서명란 ────────────────────────────────────────────
    section_title(row, '7. 확인 서명')
    row += 1
    ws.row_dimensions[row].height = 16
    ws.row_dimensions[row+1].height = 50
    merge(row, 1, row, 2);   th(row, 1, '의뢰자')
    merge(row, 3, row, 4);   th(row, 3, '부서 승인자')
    merge(row, 5, row, 6);   th(row, 5, 'QA 승인자')
    merge(row+1, 1, row+1, 2)
    merge(row+1, 3, row+1, 4)
    merge(row+1, 5, row+1, 6)
    for col in [1, 3, 5]:
        c = ws.cell(row=row+1, column=col, value='')
        c.border = b_all
    row += 2

    ws.row_dimensions[row].height = 6
    row += 1

    # ── 푸터 ─────────────────────────────────────────────────
    merge(row, 1, row, 6)
    c = ws.cell(row=row, column=1,
                value=f'본 의뢰서는 RTMS(신뢰성 시험 관리 시스템)에서 자동 생성되었습니다. '
                      f'발행: {dt.now().strftime("%Y-%m-%d %H:%M")}')
    c.font = Font(name='맑은 고딕', size=8, color='94A3B8', italic=True)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = 14

    # 인쇄 설정
    ws.page_setup.paperSize   = ws.PAPERSIZE_A4
    ws.page_setup.orientation = 'portrait'
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left   = 0.5
    ws.page_margins.right  = 0.5
    ws.page_margins.top    = 0.7
    ws.page_margins.bottom = 0.7

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    fname = f'신뢰성시험의뢰서_{req_obj.request_no or rid}_{dt.now().strftime("%Y%m%d")}.xlsx'
    return send_file(out, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ── 내부 헬퍼 ───────────────────────────────────────────────
def _save_test_items(request_id, form):
    """시험 항목 목록 저장 (form에서 배열 파라미터 읽기)"""
    names  = form.getlist('item_test_name')
    conds  = form.getlist('item_test_condition')
    stds   = form.getlist('item_standard')
    for i, (n, c, s) in enumerate(zip(names, conds, stds), 1):
        if n.strip() or c.strip() or s.strip():
            db.session.add(TestItem(
                request_id=request_id, item_no=i,
                test_name=n.strip(), test_condition=c.strip(), standard=s.strip(),
            ))


def _save_criteria(request_id, form):
    """판정 기준 목록 저장"""
    types    = form.getlist('crit_type')
    contents = form.getlist('crit_content')
    for i, (t, c) in enumerate(zip(types, contents), 1):
        if t.strip() or c.strip():
            db.session.add(TestCriterion(
                request_id=request_id, criterion_no=i,
                criterion_type=t.strip(), criterion_content=c.strip(),
            ))
