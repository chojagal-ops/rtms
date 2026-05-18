# routes/results.py — 신뢰성 시험 결과서 CRUD + 엑셀 파싱

import os, tempfile
from io import BytesIO
from werkzeug.utils import secure_filename
from flask import (Blueprint, render_template, request, redirect,
                   url_for, flash, current_app, send_from_directory, jsonify, send_file)
from flask_login import login_required, current_user
from models import db, TestRequest, TestItem, TestResult, ResultPhoto
from utils import parse_date, log_error, mail_result_notify

results_bp = Blueprint('results', __name__)

UPLOAD_FOLDER       = 'static/uploads/results'
PHOTO_FOLDER        = 'static/uploads/result_photos'
ALLOWED_EXT         = {'pdf', 'png', 'jpg', 'jpeg', 'xlsx', 'xls', 'docx', 'doc', 'zip', 'hwp'}
ALLOWED_PHOTO_EXT   = {'png', 'jpg', 'jpeg', 'gif', 'webp', 'bmp'}


def _allowed(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXT


def _save_file(file):
    if not file or file.filename == '':
        return None
    if not _allowed(file.filename):
        flash('허용되지 않는 파일 형식입니다.', 'warning')
        return None
    from datetime import datetime
    upload_dir = os.path.join(current_app.root_path, UPLOAD_FOLDER)
    os.makedirs(upload_dir, exist_ok=True)
    prefix = datetime.now().strftime('%Y%m%d%H%M%S_')
    fname  = prefix + secure_filename(file.filename)
    file.save(os.path.join(upload_dir, fname))
    return fname


def _save_photo(file):
    """결과 사진 저장 → 파일명 반환"""
    if not file or file.filename == '':
        return None
    ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else ''
    if ext not in ALLOWED_PHOTO_EXT:
        return None
    from datetime import datetime
    photo_dir = os.path.join(current_app.root_path, PHOTO_FOLDER)
    os.makedirs(photo_dir, exist_ok=True)
    import uuid
    fname = f"{datetime.now().strftime('%Y%m%d%H%M%S')}_{uuid.uuid4().hex[:8]}.{ext}"
    file.save(os.path.join(photo_dir, fname))
    return fname


def _parse_excel_result(filepath):
    """결과서 엑셀 파싱 → dict 반환"""
    import openpyxl
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active
    except Exception:
        return None

    def cell(coord):
        v = ws[coord].value
        return str(v).strip() if v is not None else ''

    data = {
        'result_date':    cell('C4') or cell('D4'),
        'complete_date':  cell('F4'),
        'overall_result': '',
        'tester_name':    cell('F27'),
        'attach_doc_name': cell('C31'),
        # 시험 결과 항목 (행 23~25)
        'item_results': [],
    }
    for row_idx in range(23, 26):
        res    = cell(f'C{row_idx}')
        detail = cell(f'D{row_idx}')
        if res or detail:
            data['item_results'].append({'result': res, 'detail': detail})
    return data


def _form_to_result(res_obj, form, files=None):
    res_obj.result_date        = parse_date(form.get('result_date'))
    res_obj.complete_date      = parse_date(form.get('complete_date'))
    res_obj.test_complete_date = parse_date(form.get('test_complete_date'))
    res_obj.notify_date        = parse_date(form.get('notify_date'))
    res_obj.overall_result     = form.get('overall_result', '').strip()
    res_obj.summary            = form.get('summary', '').strip()
    res_obj.sample_returned    = form.get('sample_returned') == 'on'
    res_obj.tester_name        = form.get('tester_name', '').strip()
    res_obj.notifier_name      = form.get('notifier_name', '').strip()
    res_obj.qa_approver        = form.get('qa_approver', '').strip()
    res_obj.report_attached    = form.get('report_attached') == 'on'
    res_obj.attach_doc_name    = form.get('attach_doc_name', '').strip()
    res_obj.notes              = form.get('notes', '').strip()
    if files:
        saved = _save_file(files.get('attachment'))
        if saved:
            res_obj.attachment = saved


# ── 결과서 목록 ─────────────────────────────────────────────
@results_bp.route('/results')
@login_required
def list_view():
    status_filter = request.args.get('status', '')
    result_filter = request.args.get('result', '')
    dept_filter   = request.args.get('dept', '')
    try:
        q = TestRequest.query
        if status_filter:
            q = q.filter(TestRequest.status == status_filter)
        else:
            q = q.filter(TestRequest.status.in_(['접수완료', '시험중', '결과회신', '완료']))
        if result_filter:
            q = q.join(TestResult, TestRequest.id == TestResult.request_id)
            q = q.filter(TestResult.overall_result == result_filter)
        if dept_filter:
            q = q.filter(TestRequest.request_dept == dept_filter)
        items = q.order_by(TestRequest.created_at.desc()).all()
        depts = [r[0] for r in db.session.query(TestRequest.request_dept)
                 .filter(TestRequest.request_dept.isnot(None),
                         TestRequest.request_dept != '',
                         TestRequest.status.in_(['접수완료', '시험중', '결과회신', '완료']))
                 .distinct().order_by(TestRequest.request_dept).all()]
    except Exception as e:
        log_error('결과서 목록 조회 오류', e)
        items = []
        depts = []
    return render_template('results/list.html', items=items,
                           status_filter=status_filter, result_filter=result_filter,
                           dept_filter=dept_filter, depts=depts)


# ── 결과서 등록/수정 ────────────────────────────────────────
@results_bp.route('/requests/<int:rid>/result/edit', methods=['GET', 'POST'])
@login_required
def edit(rid):
    """결과서 등록(신규) 또는 수정"""
    req_obj = db.get_or_404(TestRequest, rid)
    res_obj = req_obj.result   # None 이면 신규

    if request.method == 'POST':
        try:
            if res_obj is None:
                res_obj = TestResult(request_id=rid, created_by=current_user.id)
                db.session.add(res_obj)

            _form_to_result(res_obj, request.form, request.files)

            # 결과 사진 저장 (다중)
            db.session.flush()  # res_obj.id 확보
            photos      = request.files.getlist('result_photos')
            captions    = request.form.getlist('photo_captions')
            for i, photo_file in enumerate(photos):
                fname = _save_photo(photo_file)
                if fname:
                    caption = captions[i] if i < len(captions) else ''
                    db.session.add(ResultPhoto(
                        result_id=res_obj.id,
                        filename=fname,
                        caption=caption,
                    ))

            # 시험 항목별 결과 업데이트
            item_results = request.form.getlist('item_result')
            item_details = request.form.getlist('item_result_detail')
            for i, item in enumerate(req_obj.test_items):
                if i < len(item_results):
                    item.item_result  = item_results[i].strip()
                if i < len(item_details):
                    item.result_detail = item_details[i].strip()

            # 의뢰서 상태를 결과 회신으로 업데이트
            if res_obj.overall_result in ('적합', '부적합', '조건부적합'):
                req_obj.status = '결과회신'

            db.session.commit()
            flash('결과서가 저장되었습니다.', 'success')

            # 의뢰자 결과 통보 메일 발송 (비동기, overall_result 있을 때만)
            if res_obj.overall_result in ('적합', '부적합', '조건부적합'):
                try:
                    from models import User
                    creator = db.session.get(User, req_obj.created_by)
                    requester_email = creator.email if creator and creator.email else None
                    if requester_email:
                        base_url = request.url_root.rstrip('/')
                        mail_result_notify(req_obj, res_obj, requester_email, base_url)
                    else:
                        log_error('결과 통보 메일', Exception(f'의뢰자 이메일 없음 (user_id={req_obj.created_by})'))
                except Exception as e:
                    log_error('결과 통보 메일 발송 오류', e)

            return redirect(url_for('requests.detail', rid=rid))
        except Exception as e:
            db.session.rollback()
            log_error('결과서 저장 오류', e)
            flash('저장 중 오류가 발생했습니다.', 'danger')

    return render_template('results/form.html', req=req_obj, res=res_obj)


# ── 결과서 엑셀 파싱 ────────────────────────────────────────
@results_bp.route('/requests/<int:rid>/result/parse-excel', methods=['POST'])
@login_required
def parse_excel(rid):
    file = request.files.get('excel_file')
    if not file or file.filename == '':
        return jsonify({'ok': False, 'msg': '파일을 선택해주세요.'})
    ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else ''
    if ext not in ('xlsx', 'xls'):
        return jsonify({'ok': False, 'msg': 'xlsx / xls 파일만 지원합니다.'})

    with tempfile.NamedTemporaryFile(suffix=f'.{ext}', delete=False) as tmp:
        file.save(tmp.name)
        tmp_path = tmp.name
    try:
        data = _parse_excel_result(tmp_path)
        if data is None:
            return jsonify({'ok': False, 'msg': '파일을 읽을 수 없습니다.'})
        return jsonify({'ok': True, 'data': data})
    except Exception as e:
        log_error('결과서 엑셀 파싱 오류', e)
        return jsonify({'ok': False, 'msg': '파싱 중 오류가 발생했습니다.'})
    finally:
        try:
            os.remove(tmp_path)
        except Exception:
            pass


# ── 결과서 삭제 ─────────────────────────────────────────────
@results_bp.route('/requests/<int:rid>/result/delete', methods=['POST'])
@login_required
def delete(rid):
    req_obj = db.get_or_404(TestRequest, rid)
    if req_obj.result:
        try:
            db.session.delete(req_obj.result)
            req_obj.status = '시험중'
            db.session.commit()
            flash('결과서가 삭제되었습니다.', 'success')
        except Exception as e:
            db.session.rollback()
            log_error('결과서 삭제 오류', e)
            flash('삭제 중 오류가 발생했습니다.', 'danger')
    return redirect(url_for('requests.detail', rid=rid))


_RES_TEMPLATE_NET  = (r'\\192.168.10.3\품질팀\▣ SEC 무선'
                      r'\신뢰성 시험\신뢰성 시험 결과서 회신'
                      r'\신뢰성_시험_결과서_Rev.0_251113.xlsx')
# Render.com 등 클라우드 환경: static/excel_templates/ 에 파일을 커밋해두면 사용
_RES_TEMPLATE_LOCAL = os.path.join(
    os.path.dirname(os.path.dirname(__file__)),
    'static', 'excel_templates', '신뢰성_시험_결과서_Rev.0_251113.xlsx'
)


def _res_template_path():
    """네트워크 공유 → 로컬 번들 순서로 템플릿 경로 반환. 없으면 None."""
    for p in (_RES_TEMPLATE_NET, _RES_TEMPLATE_LOCAL):
        if os.path.isfile(p):
            return p
    return None


# ── 결과서 엑셀 내보내기 (원본 양식 기반) ──────────────────
@results_bp.route('/requests/<int:rid>/result/export-excel')
@login_required
def export_excel(rid):
    import openpyxl
    from openpyxl.styles import Alignment
    from datetime import datetime as dt

    req_obj = db.get_or_404(TestRequest, rid)
    res_obj = req_obj.result

    tpl = _res_template_path()
    if not tpl:
        flash('템플릿 파일을 찾을 수 없습니다. '
              'static/excel_templates/ 폴더에 결과서 양식 파일을 넣어주세요.', 'danger')
        return redirect(url_for('requests.detail', rid=rid))
    try:
        wb = openpyxl.load_workbook(tpl)
    except Exception as e:
        log_error('결과서 템플릿 로드 실패', e)
        flash('템플릿 파일을 불러올 수 없습니다.', 'danger')
        return redirect(url_for('requests.detail', rid=rid))

    ws = wb.active
    lw = Alignment(horizontal='left', vertical='center', wrap_text=True)

    def w(addr, val):
        c = ws[addr]
        c.value = str(val) if val else ''
        c.alignment = lw

    fd = lambda d: d.strftime('%Y-%m-%d') if d else ''

    # ── 의뢰서 참조 정보 (결과서 상단)
    w('C3', req_obj.request_no)
    w('C4', fd(res_obj.result_date) if res_obj else '')
    w('F4', fd(res_obj.complete_date) if res_obj else '')
    w('C6', req_obj.request_dept)
    w('C7', f"{req_obj.requester_name or ''} / {getattr(req_obj, 'requester_position', '') or ''}")
    w('C9', req_obj.contact)
    w('C10', req_obj.project_customer)
    w('F6', req_obj.product_name)
    w('F7', req_obj.model_code)
    w('F8', req_obj.product_type)
    w('F9', req_obj.test_stage)
    w('F10', req_obj.change_content)
    w('C12', req_obj.test_purpose)
    w('C13', req_obj.test_purpose_detail)
    w('F12', (str(req_obj.sample_qty) + '개') if req_obj.sample_qty else '')
    w('F13', req_obj.sample_state)
    w('F14', req_obj.sample_id_method)
    w('F15', req_obj.sample_notes)

    # ── 시험 항목 (결과서 템플릿은 단일 대형 셀 row 18)
    items = req_obj.test_items
    names_str = '\n'.join(ti.test_name or '' for ti in items)
    conds_str = '\n'.join(ti.test_condition or '' for ti in items)
    stds_str  = '\n'.join(ti.standard or '' for ti in items)
    w('C18', names_str)
    w('D18', conds_str)
    w('F18', stds_str)

    # ── 판정 기준 (row 21 단일 대형 셀)
    crits_str = '\n'.join(
        f"{cr.criterion_type or ''}: {cr.criterion_content or ''}"
        for cr in req_obj.criteria
    )
    w('C21', crits_str)

    # ── 시험 결과 (rows 24~25, 항목별)
    if res_obj:
        # row 24: 종합판정 + 첫 번째 항목 결과
        overall = res_obj.overall_result or ''
        item0_result = items[0].item_result if items else ''
        item0_detail = items[0].result_detail if items else ''
        w('C24', f"종합: {overall}\n{item0_result}" if overall else item0_result)
        w('D24', f"{res_obj.summary or ''}\n{item0_detail}" if res_obj.summary else item0_detail)

        # row 25: 두 번째 항목 이후
        remaining_results = '\n'.join(
            f"{ti.test_name or ''}: {ti.item_result or ''}" for ti in items[1:]
        )
        remaining_details = '\n'.join(ti.result_detail or '' for ti in items[1:])
        w('C25', remaining_results)
        w('D25', remaining_details)

        # ── 완료/통보 정보 (row 26)
        w('C26', f"시험완료: {fd(res_obj.test_complete_date)}\n통보: {fd(res_obj.notify_date)}")

        # ── 시료반환 / 시험자·통보자 (row 27)
        w('C27', 'YES' if res_obj.sample_returned else 'NO')
        w('F27', f"{res_obj.tester_name or ''} / {res_obj.notifier_name or ''}")

        # ── 성적서 첨부 / 첨부문서 (rows 29, 31)
        w('C29', 'YES' if res_obj.report_attached else 'NO')
        w('C31', res_obj.attach_doc_name or '')

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    fname = f'신뢰성_시험_결과서_{req_obj.request_no or rid}_{dt.now().strftime("%Y%m%d")}.xlsx'
    return send_file(out, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ── 파일 다운로드 ────────────────────────────────────────────
@results_bp.route('/results/uploads/<filename>')
@login_required
def download_file(filename):
    upload_dir = os.path.join(current_app.root_path, UPLOAD_FOLDER)
    return send_from_directory(upload_dir, filename, as_attachment=True)


# ── 결과 사진 서빙 ──────────────────────────────────────────
@results_bp.route('/results/photos/<filename>')
@login_required
def serve_photo(filename):
    photo_dir = os.path.join(current_app.root_path, PHOTO_FOLDER)
    return send_from_directory(photo_dir, filename)


# ── 결과 사진 삭제 ──────────────────────────────────────────
@results_bp.route('/results/photos/<int:pid>/delete', methods=['POST'])
@login_required
def delete_photo(pid):
    photo = db.get_or_404(ResultPhoto, pid)
    rid   = photo.result.request_id
    try:
        path = os.path.join(current_app.root_path, PHOTO_FOLDER, photo.filename)
        if os.path.exists(path):
            os.remove(path)
        db.session.delete(photo)
        db.session.commit()
        flash('사진이 삭제되었습니다.', 'success')
    except Exception as e:
        db.session.rollback()
        log_error('결과 사진 삭제 오류', e)
        flash('삭제 중 오류가 발생했습니다.', 'danger')
    return redirect(url_for('requests.detail', rid=rid))


# ── 성적서 인쇄 뷰 ───────────────────────────────────────────
@results_bp.route('/requests/<int:rid>/result-certificate')
@login_required
def certificate(rid):
    """성적서 인쇄용 HTML 뷰"""
    req_obj = db.get_or_404(TestRequest, rid)
    res_obj = req_obj.result
    if not res_obj:
        flash('결과서가 아직 등록되지 않았습니다.', 'warning')
        return redirect(url_for('requests.detail', rid=rid))
    return render_template('results/certificate.html', req=req_obj, res=res_obj)


# ── 성적서 엑셀 다운로드 ─────────────────────────────────────
@results_bp.route('/requests/<int:rid>/result-certificate/excel')
@login_required
def certificate_excel(rid):
    """성적서 엑셀 생성 (openpyxl 직접 생성)"""
    import openpyxl
    from openpyxl.styles import (Font, Alignment, PatternFill,
                                 Border, Side, numbers)
    from openpyxl.utils import get_column_letter
    from datetime import datetime as dt

    req_obj = db.get_or_404(TestRequest, rid)
    res_obj = req_obj.result
    if not res_obj:
        flash('결과서가 아직 등록되지 않았습니다.', 'warning')
        return redirect(url_for('requests.detail', rid=rid))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '시험성적서'

    # 열 너비
    col_widths = [6, 18, 32, 16, 12, 28]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    thin  = Side(style='thin',   color='000000')
    thick = Side(style='medium', color='000000')
    b_all = Border(left=thin, right=thin, top=thin, bottom=thin)
    b_thick = Border(left=thick, right=thick, top=thick, bottom=thick)

    def cell(row, col, value='', bold=False, size=10, align='left',
             valign='center', fill=None, border=None, wrap=False):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(name='맑은 고딕', bold=bold, size=size)
        c.alignment = Alignment(horizontal=align, vertical=valign,
                                wrap_text=wrap)
        if fill:
            c.fill = PatternFill('solid', fgColor=fill)
        if border:
            c.border = border
        return c

    def merge(r1, c1, r2, c2):
        ws.merge_cells(start_row=r1, start_column=c1,
                       end_row=r2, end_column=c2)

    fd = lambda d: d.strftime('%Y-%m-%d') if d else '-'
    ORANGE = 'F97316'
    LGRAY  = 'F1F5F9'
    DGRAY  = '475569'

    # ── 제목 헤더 ──────────────────────────────────────────
    ws.row_dimensions[1].height = 14
    ws.row_dimensions[2].height = 36
    ws.row_dimensions[3].height = 14

    merge(2, 1, 2, 6)
    cell(2, 1, '신뢰성 시험 성적서', bold=True, size=18, align='center',
         fill=ORANGE)
    ws.cell(2, 1).font = Font(name='맑은 고딕', bold=True, size=18,
                               color='FFFFFF')

    merge(4, 1, 4, 3)
    cell(4, 1, f'성적서번호: {req_obj.request_no or "-"}', bold=True, size=10)
    merge(4, 4, 4, 6)
    cell(4, 4, f'발행일: {fd(res_obj.result_date or res_obj.created_at.date() if res_obj.created_at else None)}',
         align='right', size=10)

    ws.row_dimensions[5].height = 6

    # ── 섹션1: 의뢰 정보 ────────────────────────────────────
    merge(6, 1, 6, 6)
    c = cell(6, 1, '1. 시험 의뢰 정보', bold=True, size=10,
             fill=LGRAY, border=b_all)
    ws.cell(6, 1).font = Font(name='맑은 고딕', bold=True, size=10,
                               color=DGRAY)

    headers1 = ['의뢰번호', '의뢰부서', '의뢰자', '제품명', '모델명/코드', '의뢰일자']
    values1  = [
        req_obj.request_no or '-',
        req_obj.request_dept or '-',
        f"{req_obj.requester_name or '-'} {req_obj.requester_position or ''}".strip(),
        req_obj.product_name or '-',
        req_obj.model_code or '-',
        fd(req_obj.write_date),
    ]
    for col, (h, v) in enumerate(zip(headers1, values1), 1):
        ws.row_dimensions[7].height = 16
        ws.row_dimensions[8].height = 20
        c = ws.cell(row=7, column=col, value=h)
        c.font = Font(name='맑은 고딕', bold=True, size=9, color='FFFFFF')
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.fill = PatternFill('solid', fgColor='334155')
        c.border = b_all
        cv = ws.cell(row=8, column=col, value=v)
        cv.font = Font(name='맑은 고딕', size=10,
                       bold=(col in (4,)))
        cv.alignment = Alignment(horizontal='center', vertical='center',
                                 wrap_text=True)
        cv.border = b_all

    headers2 = ['제품 종류', '시험 단계', '시험 목적', '시편 수량', '완료희망일', '우선순위']
    values2  = [
        req_obj.product_type or '-',
        req_obj.test_stage or '-',
        req_obj.test_purpose or '-',
        f"{req_obj.sample_qty or '-'}개" if req_obj.sample_qty else '-',
        fd(req_obj.deadline),
        req_obj.priority or '-',
    ]
    for col, (h, v) in enumerate(zip(headers2, values2), 1):
        ws.row_dimensions[9].height = 16
        ws.row_dimensions[10].height = 18
        c = ws.cell(row=9, column=col, value=h)
        c.font = Font(name='맑은 고딕', bold=True, size=9, color='FFFFFF')
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.fill = PatternFill('solid', fgColor='334155')
        c.border = b_all
        cv = ws.cell(row=10, column=col, value=v)
        cv.font = Font(name='맑은 고딕', size=10)
        cv.alignment = Alignment(horizontal='center', vertical='center')
        cv.border = b_all

    ws.row_dimensions[11].height = 6

    # ── 섹션2: 시험 항목별 결과 ──────────────────────────────
    merge(12, 1, 12, 6)
    cell(12, 1, '2. 시험 항목별 결과', bold=True, size=10,
         fill=LGRAY, border=b_all)
    ws.cell(12, 1).font = Font(name='맑은 고딕', bold=True, size=10,
                                color=DGRAY)

    item_headers = ['No', '시험 항목', '시험 조건', '규격', '판정', '결과 상세']
    ws.row_dimensions[13].height = 16
    for col, h in enumerate(item_headers, 1):
        c = ws.cell(row=13, column=col, value=h)
        c.font = Font(name='맑은 고딕', bold=True, size=9, color='FFFFFF')
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.fill = PatternFill('solid', fgColor='334155')
        c.border = b_all

    row = 14
    items = req_obj.test_items
    for ti in items:
        ws.row_dimensions[row].height = 28
        result_val = ti.item_result or '-'
        result_color = ('15803d' if result_val == '적합' else
                        'dc2626' if result_val == '부적합' else '374151')
        vals = [ti.item_no, ti.test_name or '-', ti.test_condition or '-',
                ti.standard or '-', result_val, ti.result_detail or '-']
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=v)
            c.font = Font(name='맑은 고딕', size=9,
                          bold=(col == 5),
                          color=(result_color if col == 5 else '000000'))
            c.alignment = Alignment(horizontal='center' if col in (1,4,5) else 'left',
                                    vertical='center', wrap_text=True)
            c.border = b_all
        row += 1

    if not items:
        merge(14, 1, 14, 6)
        cell(14, 1, '등록된 시험 항목이 없습니다.', align='center', size=9)
        ws.cell(14, 1).border = b_all
        row = 15

    ws.row_dimensions[row].height = 6
    row += 1

    # ── 섹션3: 종합 판정 ─────────────────────────────────────
    merge(row, 1, row, 6)
    cell(row, 1, '3. 종합 판정', bold=True, size=10, fill=LGRAY, border=b_all)
    ws.cell(row, 1).font = Font(name='맑은 고딕', bold=True, size=10,
                                 color=DGRAY)
    row += 1

    overall = res_obj.overall_result or '-'
    ov_color = ('15803d' if overall == '적합' else
                'dc2626' if overall == '부적합' else
                'd97706' if overall == '조건부적합' else '374151')

    ws.row_dimensions[row].height = 18
    ws.row_dimensions[row+1].height = 50
    for col, h in enumerate(['종합 판정 결과', '결과 요약 / 종합 의견'], 1):
        c = ws.cell(row=row, column=col if col == 1 else 2, value=h)
        c.font = Font(name='맑은 고딕', bold=True, size=9, color='FFFFFF')
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.fill = PatternFill('solid', fgColor='334155')
        c.border = b_all
    merge(row, 2, row, 6)

    merge(row+1, 1, row+1, 1)
    c = ws.cell(row=row+1, column=1, value=overall)
    c.font = Font(name='맑은 고딕', bold=True, size=16, color=ov_color)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = b_all

    merge(row+1, 2, row+1, 6)
    c = ws.cell(row=row+1, column=2, value=res_obj.summary or '')
    c.font = Font(name='맑은 고딕', size=10)
    c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    c.border = b_all
    row += 2

    ws.row_dimensions[row].height = 6
    row += 1

    # ── 섹션4: 시험 완료 정보 ────────────────────────────────
    merge(row, 1, row, 6)
    cell(row, 1, '4. 시험 완료 정보', bold=True, size=10,
         fill=LGRAY, border=b_all)
    ws.cell(row, 1).font = Font(name='맑은 고딕', bold=True, size=10,
                                 color=DGRAY)
    row += 1

    info_headers = ['시험 완료일', '통보일', '시험자', '통보자', '품질팀 승인자', '시료 반환']
    info_values  = [
        fd(res_obj.test_complete_date),
        fd(res_obj.notify_date),
        res_obj.tester_name or '-',
        res_obj.notifier_name or '-',
        res_obj.qa_approver or req_obj.qa_approver or '-',
        'YES' if res_obj.sample_returned else 'NO',
    ]
    ws.row_dimensions[row].height = 16
    ws.row_dimensions[row+1].height = 20
    for col, (h, v) in enumerate(zip(info_headers, info_values), 1):
        ch = ws.cell(row=row, column=col, value=h)
        ch.font = Font(name='맑은 고딕', bold=True, size=9, color='FFFFFF')
        ch.alignment = Alignment(horizontal='center', vertical='center')
        ch.fill = PatternFill('solid', fgColor='334155')
        ch.border = b_all
        cv = ws.cell(row=row+1, column=col, value=v)
        cv.font = Font(name='맑은 고딕', size=10)
        cv.alignment = Alignment(horizontal='center', vertical='center')
        cv.border = b_all
    row += 2

    ws.row_dimensions[row].height = 6
    row += 1

    # ── 섹션5: 서명란 ────────────────────────────────────────
    merge(row, 1, row, 6)
    cell(row, 1, '5. 확인 서명', bold=True, size=10, fill=LGRAY, border=b_all)
    ws.cell(row, 1).font = Font(name='맑은 고딕', bold=True, size=10, color=DGRAY)
    row += 1

    sign_headers = ['시험자', '', 'QA 담당자', '', 'QA 승인자', '']
    ws.row_dimensions[row].height = 16
    ws.row_dimensions[row+1].height = 40
    for col, h in enumerate(sign_headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = Font(name='맑은 고딕', bold=True, size=9, color='FFFFFF')
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.fill = PatternFill('solid', fgColor='334155')
        c.border = b_all
        cv = ws.cell(row=row+1, column=col, value='')
        cv.border = b_all

    merge(row, 1, row, 2)
    merge(row, 3, row, 4)
    merge(row, 5, row, 6)
    merge(row+1, 1, row+1, 2)
    merge(row+1, 3, row+1, 4)
    merge(row+1, 5, row+1, 6)

    sign_labels = ['시험자', 'QA 담당자', '품질팀 승인자']
    for i, label in enumerate(sign_labels):
        col = 1 + i * 2
        c = ws.cell(row=row, column=col, value=label)
        c.font = Font(name='맑은 고딕', bold=True, size=9, color='FFFFFF')
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.fill = PatternFill('solid', fgColor='334155')
        c.border = b_all

    row += 2
    ws.row_dimensions[row].height = 6
    row += 1

    # ── 푸터 ─────────────────────────────────────────────────
    merge(row, 1, row, 6)
    c = ws.cell(row=row, column=1,
                value=f'본 성적서는 RTMS(신뢰성 시험 관리 시스템)에서 자동 생성되었습니다. '
                      f'발행: {dt.now().strftime("%Y-%m-%d %H:%M")}')
    c.font = Font(name='맑은 고딕', size=8, color='94A3B8', italic=True)
    c.alignment = Alignment(horizontal='center', vertical='center')

    # 인쇄 설정
    ws.page_setup.paperSize  = ws.PAPERSIZE_A4
    ws.page_setup.orientation = 'portrait'
    ws.page_setup.fitToPage  = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left   = 0.5
    ws.page_margins.right  = 0.5
    ws.page_margins.top    = 0.7
    ws.page_margins.bottom = 0.7

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    fname = f'신뢰성시험성적서_{req_obj.request_no or rid}_{dt.now().strftime("%Y%m%d")}.xlsx'
    return send_file(out, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
