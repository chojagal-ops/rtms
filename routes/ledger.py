# routes/ledger.py — 신뢰성 시험 관리대장

from io import BytesIO
from flask import Blueprint, render_template, request, send_file
from flask_login import login_required
from datetime import date, datetime
from models import db, TestRequest
from utils import log_error
from constants import REQUEST_STATUSES

ledger_bp = Blueprint('ledger', __name__)


@ledger_bp.route('/ledger')
@login_required
def index():
    year   = request.args.get('year',   '')
    month  = request.args.get('month',  '')
    dept   = request.args.get('dept',   '').strip()
    status = request.args.get('status', '')
    q_str  = request.args.get('q',      '').strip()

    try:
        q = TestRequest.query

        if year:
            q = q.filter(db.func.strftime('%Y', TestRequest.write_date) == year.zfill(4))
        if month:
            q = q.filter(db.func.strftime('%m', TestRequest.write_date) == month.zfill(2))
        if dept:
            q = q.filter(TestRequest.request_dept.ilike(f'%{dept}%'))
        if status:
            q = q.filter(TestRequest.status == status)
        if q_str:
            q = q.filter(db.or_(
                TestRequest.request_no.ilike(f'%{q_str}%'),
                TestRequest.product_name.ilike(f'%{q_str}%'),
                TestRequest.requester_name.ilike(f'%{q_str}%'),
            ))

        items = q.order_by(
            TestRequest.write_date.desc(),
            TestRequest.id.desc()
        ).all()

        # 부서 목록 (필터용)
        dept_rows = (db.session.query(TestRequest.request_dept)
                     .distinct()
                     .filter(TestRequest.request_dept.isnot(None))
                     .all())
        depts = sorted([d[0] for d in dept_rows if d[0]])

        current_year = date.today().year
        years = list(range(current_year, current_year - 6, -1))

    except Exception as e:
        log_error('관리대장 조회 오류', e)
        items, depts, years = [], [], []

    return render_template(
        'ledger.html',
        items=items, depts=depts, years=years,
        year=year, month=month, dept=dept, status=status, q_str=q_str,
        today=date.today(),
        REQUEST_STATUSES=REQUEST_STATUSES,
    )


def _build_ledger_query(args):
    year   = args.get('year',   '')
    month  = args.get('month',  '')
    dept   = args.get('dept',   '').strip()
    status = args.get('status', '')
    q_str  = args.get('q',      '').strip()
    q = TestRequest.query
    if year:   q = q.filter(db.func.strftime('%Y', TestRequest.write_date) == year.zfill(4))
    if month:  q = q.filter(db.func.strftime('%m', TestRequest.write_date) == month.zfill(2))
    if dept:   q = q.filter(TestRequest.request_dept.ilike(f'%{dept}%'))
    if status: q = q.filter(TestRequest.status == status)
    if q_str:  q = q.filter(db.or_(
        TestRequest.request_no.ilike(f'%{q_str}%'),
        TestRequest.product_name.ilike(f'%{q_str}%'),
        TestRequest.requester_name.ilike(f'%{q_str}%'),
    ))
    return q.order_by(TestRequest.write_date.desc(), TestRequest.id.desc()).all()


@ledger_bp.route('/ledger/export-excel')
@login_required
def export_excel():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    items = _build_ledger_query(request.args)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '관리대장'

    thin = Side(style='thin')
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill('solid', fgColor='1A3A5C')
    hdr_font = Font(color='FFFFFF', bold=True, size=10)
    center   = Alignment(horizontal='center', vertical='center', wrap_text=True)
    middle   = Alignment(vertical='center', wrap_text=True)

    # 제목
    ws.merge_cells('A1:P1')
    ws['A1'] = '신뢰성 시험 관리대장'
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = center
    ws.row_dimensions[1].height = 28

    # 헤더
    headers = ['No.','의뢰일자','의뢰부서','의뢰자','품명/모델명','LOT No./S/N',
               '시험 의뢰 목적','시험 항목','시험 조건','시험 시작일','시험 완료일',
               '시험자','시험 결과','판정','결과 통보일','특이사항/조치사항']
    ws.append(headers)
    for cell in ws[2]:
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = center
        cell.border = bdr
    ws.row_dimensions[2].height = 22

    # 데이터
    for idx, item in enumerate(items, 1):
        res = item.result
        names = '\n'.join(ti.test_name or '' for ti in item.test_items if ti.test_name)
        conds = '\n'.join(ti.test_condition or '' for ti in item.test_items if ti.test_condition)
        prod  = item.product_name or ''
        if item.model_code:
            prod += f'\n{item.model_code}'
        row = [
            idx,
            item.write_date.strftime('%Y-%m-%d') if item.write_date else '',
            item.request_dept or '',
            item.requester_name or '',
            prod,
            item.sample_id_method or '',
            item.test_purpose or '',
            names,
            conds,
            item.req_start_date.strftime('%Y-%m-%d') if item.req_start_date else '',
            res.test_complete_date.strftime('%Y-%m-%d') if res and res.test_complete_date else '',
            res.tester_name if res else '',
            res.summary or (res.overall_result or '') if res else '',
            res.overall_result if res else '',
            res.notify_date.strftime('%Y-%m-%d') if res and res.notify_date else '',
            (res.notes or '') or (item.notes or '') if (res or item.notes) else '',
        ]
        ws.append(row)
        r = ws.max_row
        for cell in ws[r]:
            cell.border = bdr
            cell.alignment = middle
        ws.row_dimensions[r].height = 18

    # 컬럼 폭
    col_widths = [5,12,12,10,18,14,12,18,22,12,12,10,30,10,12,24]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    ws.freeze_panes = 'A3'

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    fname = f'신뢰성_시험_관리대장_{datetime.now().strftime("%Y%m%d")}.xlsx'
    return send_file(out, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
