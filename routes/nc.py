# routes/nc.py — 부적합 관리

import os
from datetime import date, datetime
from werkzeug.utils import secure_filename
from flask import (Blueprint, render_template, request, redirect,
                   url_for, flash, current_app, send_from_directory)
from flask_login import login_required, current_user
from models import db, NCReport, NCAction, NCPhoto, TestRequest
from utils import parse_date, log_error, mail_nc_notify
from constants import NC_STATUSES, NC_SEVERITIES

NC_PHOTO_FOLDER   = 'static/uploads/nc_photos'
ALLOWED_PHOTO_EXT = {'png', 'jpg', 'jpeg', 'gif', 'webp', 'bmp'}


def _save_nc_photo(file):
    if not file or file.filename == '':
        return None
    ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else ''
    if ext not in ALLOWED_PHOTO_EXT:
        return None
    upload_dir = os.path.join(current_app.root_path, NC_PHOTO_FOLDER)
    os.makedirs(upload_dir, exist_ok=True)
    import uuid
    fname = datetime.now().strftime('%Y%m%d%H%M%S_') + uuid.uuid4().hex[:8] + '.' + ext
    file.save(os.path.join(upload_dir, fname))
    return fname

nc_bp = Blueprint('nc', __name__)


def _next_nc_no():
    """부적합 번호 채번 NC-YYYYMM-NNN"""
    today  = date.today()
    prefix = f'NC-{today.strftime("%Y%m")}'
    last   = (NCReport.query
              .filter(NCReport.nc_no.like(f'{prefix}%'))
              .order_by(NCReport.nc_no.desc()).first())
    seq = 1
    if last and last.nc_no:
        try:
            seq = int(last.nc_no.split('-')[-1]) + 1
        except (ValueError, IndexError):
            seq = 1
    return f'{prefix}-{seq:03d}'


def _form_to_nc(nc, form):
    # 시험의뢰자 정보
    nc.requester_name    = form.get('requester_name', '').strip()
    nc.requester_dept    = form.get('requester_dept', '').strip()
    nc.requester_contact = form.get('requester_contact', '').strip()
    nc.requester_email   = form.get('requester_email', '').strip()
    # 기본 정보
    nc.detection_date  = parse_date(form.get('detection_date'))
    nc.detected_by     = form.get('detected_by', '').strip()
    nc.product_name    = form.get('product_name', '').strip()
    nc.defect_type     = form.get('defect_type', '').strip()
    nc.defect_desc     = form.get('defect_desc', '').strip()
    qty = form.get('quantity', '').strip()
    nc.quantity        = int(qty) if qty.isdigit() else None
    nc.severity        = form.get('severity', '').strip()
    nc.dept            = form.get('dept', '').strip()
    nc.status          = form.get('status', nc.status or '등록')
    nc.notes           = form.get('notes', '').strip()
    rid = form.get('request_id', '').strip()
    nc.request_id      = int(rid) if rid.isdigit() else None



# ── 부적합 목록 (불량 등록) ─────────────────────────────────
@nc_bp.route('/nc')
@login_required
def list_view():
    status_f   = request.args.get('status', '')
    severity_f = request.args.get('severity', '')
    dept_f     = request.args.get('dept', '')
    search     = request.args.get('q', '').strip()
    try:
        q = NCReport.query
        if status_f:
            q = q.filter(NCReport.status == status_f)
        if severity_f:
            q = q.filter(NCReport.severity == severity_f)
        if dept_f:
            q = q.filter(NCReport.dept == dept_f)
        if search:
            q = q.filter(db.or_(
                NCReport.nc_no.ilike(f'%{search}%'),
                NCReport.product_name.ilike(f'%{search}%'),
                NCReport.defect_type.ilike(f'%{search}%'),
                NCReport.detected_by.ilike(f'%{search}%'),
            ))
        items = q.order_by(NCReport.created_at.desc()).all()
        depts = [r[0] for r in db.session.query(NCReport.dept)
                 .filter(NCReport.dept.isnot(None), NCReport.dept != '')
                 .distinct().order_by(NCReport.dept).all()]
    except Exception as e:
        log_error('부적합 목록 조회 오류', e)
        items = []
        depts = []
    return render_template('nc/list.html', items=items,
                           status_f=status_f, severity_f=severity_f,
                           dept_f=dept_f, depts=depts,
                           search=search,
                           NC_STATUSES=NC_STATUSES, NC_SEVERITIES=NC_SEVERITIES)


# ── 부적합 등록 ─────────────────────────────────────────────
@nc_bp.route('/nc/new', methods=['GET', 'POST'])
@login_required
def new():
    if request.method == 'POST':
        try:
            nc = NCReport(created_by=current_user.id,
                          nc_no=_next_nc_no(), status='등록')
            _form_to_nc(nc, request.form)
            db.session.add(nc)
            db.session.flush()
            # 사진 저장
            photos   = request.files.getlist('nc_photos')
            captions = request.form.getlist('photo_captions')
            for i, photo_file in enumerate(photos):
                fname = _save_nc_photo(photo_file)
                if fname:
                    db.session.add(NCPhoto(
                        nc_id=nc.id, filename=fname,
                        caption=captions[i] if i < len(captions) else ''))
            db.session.commit()
            try:
                base_url = request.host_url.rstrip('/')
                mail_nc_notify(nc, base_url)
            except Exception:
                pass
            flash(f'부적합 [{nc.nc_no}] 등록되었습니다.', 'success')
            return redirect(url_for('nc.detail', nid=nc.id))
        except Exception as e:
            db.session.rollback()
            log_error('부적합 등록 오류', e)
            flash('등록 중 오류가 발생했습니다.', 'danger')
    all_reqs = (TestRequest.query
                .order_by(TestRequest.created_at.desc()).limit(300).all())
    return render_template('nc/form.html', nc=None, all_reqs=all_reqs,
                           NC_STATUSES=NC_STATUSES,
                           NC_SEVERITIES=NC_SEVERITIES)


# ── 부적합 상세 ─────────────────────────────────────────────
@nc_bp.route('/nc/<int:nid>')
@login_required
def detail(nid):
    nc = db.get_or_404(NCReport, nid)
    return render_template('nc/detail.html', nc=nc, NC_STATUSES=NC_STATUSES)


# ── 부적합 수정 ─────────────────────────────────────────────
@nc_bp.route('/nc/<int:nid>/edit', methods=['GET', 'POST'])
@login_required
def edit(nid):
    nc = db.get_or_404(NCReport, nid)
    if request.method == 'POST':
        try:
            _form_to_nc(nc, request.form)
            # 추가 사진 저장
            photos   = request.files.getlist('nc_photos')
            captions = request.form.getlist('photo_captions')
            for i, photo_file in enumerate(photos):
                fname = _save_nc_photo(photo_file)
                if fname:
                    db.session.add(NCPhoto(
                        nc_id=nc.id, filename=fname,
                        caption=captions[i] if i < len(captions) else ''))
            db.session.commit()
            flash('수정되었습니다.', 'success')
            return redirect(url_for('nc.detail', nid=nid))
        except Exception as e:
            db.session.rollback()
            log_error('부적합 수정 오류', e)
            flash('수정 중 오류가 발생했습니다.', 'danger')
    all_reqs = (TestRequest.query
                .order_by(TestRequest.created_at.desc()).limit(300).all())
    return render_template('nc/form.html', nc=nc, all_reqs=all_reqs,
                           NC_STATUSES=NC_STATUSES,
                           NC_SEVERITIES=NC_SEVERITIES)



# ── NC 종료 처리 ──────────────────────────────────────────
@nc_bp.route('/nc/<int:nid>/close', methods=['POST'])
@login_required
def close_nc(nid):
    nc = db.get_or_404(NCReport, nid)
    try:
        nc.status = '종료'
        db.session.commit()
        flash(f'[{nc.nc_no}] 종료 처리되었습니다.', 'success')
    except Exception as e:
        db.session.rollback()
        log_error('NC 종료 오류', e)
        flash('처리 중 오류가 발생했습니다.', 'danger')
    return redirect(url_for('nc.detail', nid=nid))


# ── NC 삭제 ───────────────────────────────────────────────
@nc_bp.route('/nc/<int:nid>/delete', methods=['POST'])
@login_required
def delete(nid):
    nc = db.get_or_404(NCReport, nid)
    try:
        db.session.delete(nc)
        db.session.commit()
        flash('부적합 보고서가 삭제되었습니다.', 'success')
    except Exception as e:
        db.session.rollback()
        log_error('NC 삭제 오류', e)
        flash('삭제 중 오류가 발생했습니다.', 'danger')
    return redirect(url_for('nc.list_view'))


# ── NC 사진 서빙 ──────────────────────────────────────────
@nc_bp.route('/nc/photos/<path:filename>')
@login_required
def serve_photo(filename):
    photo_dir = os.path.join(current_app.root_path, NC_PHOTO_FOLDER)
    return send_from_directory(photo_dir, filename)


# ── NC 사진 삭제 ──────────────────────────────────────────
@nc_bp.route('/nc/photos/<int:pid>/delete', methods=['POST'])
@login_required
def delete_photo(pid):
    photo = db.get_or_404(NCPhoto, pid)
    nid   = photo.nc_id
    try:
        fpath = os.path.join(current_app.root_path, NC_PHOTO_FOLDER, photo.filename)
        if os.path.exists(fpath):
            os.remove(fpath)
        db.session.delete(photo)
        db.session.commit()
        flash('사진이 삭제되었습니다.', 'success')
    except Exception as e:
        db.session.rollback()
        log_error('NC 사진 삭제 오류', e)
        flash('삭제 중 오류가 발생했습니다.', 'danger')
    return redirect(url_for('nc.detail', nid=nid))


# ── 처리내용 등록 ─────────────────────────────────────────
@nc_bp.route('/nc/<int:nid>/response', methods=['POST'])
@login_required
def add_response(nid):
    nc = db.get_or_404(NCReport, nid)
    desc = request.form.get('action_desc', '').strip()
    if not desc:
        flash('처리 내용을 입력해주세요.', 'warning')
        return redirect(url_for('nc.detail', nid=nid))
    try:
        action = NCAction(
            nc_id           = nc.id,
            action_type     = '처리내용',
            action_desc     = desc,
            responsible     = request.form.get('responsible', '').strip() or current_user.name,
            completion_date = parse_date(request.form.get('completion_date')),
            status          = request.form.get('action_status', '처리중'),
        )
        db.session.add(action)
        new_nc_status = request.form.get('nc_status', '').strip()
        if new_nc_status and new_nc_status in NC_STATUSES:
            nc.status = new_nc_status
        db.session.commit()
        flash('처리내용이 등록되었습니다.', 'success')
    except Exception as e:
        db.session.rollback()
        log_error('처리내용 등록 오류', e)
        flash('등록 중 오류가 발생했습니다.', 'danger')
    return redirect(url_for('nc.detail', nid=nid))


# ── 처리내용 삭제 ─────────────────────────────────────────
@nc_bp.route('/nc/response/<int:aid>/delete', methods=['POST'])
@login_required
def delete_response(aid):
    action = db.get_or_404(NCAction, aid)
    nid = action.nc_id
    try:
        db.session.delete(action)
        db.session.commit()
        flash('처리내용이 삭제되었습니다.', 'success')
    except Exception as e:
        db.session.rollback()
        log_error('처리내용 삭제 오류', e)
        flash('삭제 중 오류가 발생했습니다.', 'danger')
    return redirect(url_for('nc.detail', nid=nid))


# ── 부적합 관리이력 ───────────────────────────────────────
@nc_bp.route('/nc/history')
@login_required
def history():
    date_from  = request.args.get('date_from', '')
    date_to    = request.args.get('date_to', '')
    status_f   = request.args.get('status', '')
    severity_f = request.args.get('severity', '')
    dept_f     = request.args.get('dept', '')
    search     = request.args.get('q', '').strip()
    q = NCReport.query
    if date_from:
        q = q.filter(NCReport.detection_date >= parse_date(date_from))
    if date_to:
        q = q.filter(NCReport.detection_date <= parse_date(date_to))
    if status_f:
        q = q.filter(NCReport.status == status_f)
    if severity_f:
        q = q.filter(NCReport.severity == severity_f)
    if dept_f:
        q = q.filter(NCReport.dept == dept_f)
    if search:
        q = q.filter(db.or_(
            NCReport.nc_no.ilike(f'%{search}%'),
            NCReport.product_name.ilike(f'%{search}%'),
            NCReport.defect_type.ilike(f'%{search}%'),
        ))
    items = q.order_by(NCReport.detection_date.desc(), NCReport.created_at.desc()).all()
    depts = [r[0] for r in db.session.query(NCReport.dept)
             .filter(NCReport.dept.isnot(None), NCReport.dept != '')
             .distinct().order_by(NCReport.dept).all()]
    return render_template('nc/history.html', items=items, depts=depts,
                           date_from=date_from, date_to=date_to,
                           status_f=status_f, severity_f=severity_f,
                           dept_f=dept_f, search=search,
                           NC_STATUSES=NC_STATUSES, NC_SEVERITIES=NC_SEVERITIES)


# ── 부적합 관리이력 엑셀 추출 ────────────────────────────
@nc_bp.route('/nc/history/export')
@login_required
def history_export():
    from io import BytesIO
    from flask import send_file
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    import openpyxl.utils

    date_from  = request.args.get('date_from', '')
    date_to    = request.args.get('date_to', '')
    status_f   = request.args.get('status', '')
    severity_f = request.args.get('severity', '')
    dept_f     = request.args.get('dept', '')
    search     = request.args.get('q', '').strip()
    q = NCReport.query
    if date_from:
        q = q.filter(NCReport.detection_date >= parse_date(date_from))
    if date_to:
        q = q.filter(NCReport.detection_date <= parse_date(date_to))
    if status_f:
        q = q.filter(NCReport.status == status_f)
    if severity_f:
        q = q.filter(NCReport.severity == severity_f)
    if dept_f:
        q = q.filter(NCReport.dept == dept_f)
    if search:
        q = q.filter(db.or_(
            NCReport.nc_no.ilike(f'%{search}%'),
            NCReport.product_name.ilike(f'%{search}%'),
            NCReport.defect_type.ilike(f'%{search}%'),
        ))
    items = q.order_by(NCReport.detection_date.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '부적합 관리이력'
    thin   = Side(style='thin', color='D1D5DB')
    bdr    = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    wrap   = Alignment(vertical='top', wrap_text=True)
    headers = ['No','NC번호','발생일','제품명/모델','부적합유형','심각도',
               '발생부서','부적합내용','상태','의뢰자','처리내용','처리자','처리일']
    col_w   = [5,16,12,22,18,8,14,35,10,18,35,12,12]
    ws.row_dimensions[1].height = 22
    for ci,(h,w) in enumerate(zip(headers,col_w),1):
        cell = ws.cell(row=1,column=ci,value=h)
        cell.font      = Font(bold=True,color='FFFFFF',size=11)
        cell.fill      = PatternFill('solid',fgColor='F97316')
        cell.alignment = center
        cell.border    = bdr
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w

    fd = lambda d: d.strftime('%Y-%m-%d') if d else ''
    for row_i,nc in enumerate(items,2):
        responses = [a for a in nc.actions if a.action_type == '처리내용']
        latest    = responses[-1] if responses else None
        req_str   = nc.requester_name or ''
        if nc.requester_dept:
            req_str += f' ({nc.requester_dept})'
        vals = [
            row_i-1, nc.nc_no or '', fd(nc.detection_date),
            nc.product_name or '', nc.defect_type or '', nc.severity or '',
            nc.dept or '', nc.defect_desc or '', nc.status or '', req_str,
            latest.action_desc if latest else '',
            latest.responsible if latest else '',
            fd(latest.completion_date) if latest else '',
        ]
        ws.row_dimensions[row_i].height = 40
        for ci,v in enumerate(vals,1):
            cell = ws.cell(row=row_i,column=ci,value=v)
            cell.border    = bdr
            cell.alignment = center if ci in (1,6,9) else wrap
            if row_i % 2 == 0:
                cell.fill = PatternFill('solid',fgColor='FFF7ED')

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f'부적합관리이력_{date.today().strftime("%Y%m%d")}.xlsx'
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

